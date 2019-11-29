#/bin/python
# -*- coding=uat-8 -*-

"""
从项目的relsenotes中抽取发布包列表
notes特点
命名规则：项目简写_Release_Notes_版本号_8位日期.docx
绝大多数notes中第四个table是发布包列表，比较早的notes是第3个table,所以通过判断Table的第一行的title 确定table的索引
部分notes中的package是放在excel中的，如果table的行数为1，说明通过excel保存的packages list
部分notes中有两个excel，第二个excel是 package，第一个是解决问题列表
部分notes中有excel中的package 只有3列，有的是2列。3列多了一列父目录，此目录名可以不用抽取。
"""
import os, sys
from docx import Document
import shutil
import zipfile
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


# path of release notes
dirOfNotes = 'D:\\17_Book\\demo'
#dirOfNotes = 'D:\\17_Book\\64_CCIC\\Release_Notes'

def info(*mes):
    print('[Info]>>>', *mes)

def err(mes):
    print('[Error]>>>', mes)

def readNotes():
    # 读取目录下下面的docx文件，并且按照时间排序
    allNotes = sorted(glob.glob(os.path.join(dirOfNotes, '*.docx')), key=os.path.getmtime)
    for notes in allNotes:
        #info(notes)
        # glob读取的文件是绝对路径，这个和os.listdir() 不一样
        # 从文件名中拆分日期
        releaseDate = os.path.splitext(os.path.basename(notes))[0].split('_')[4]
        # 从文件名中拆分版本
        noteVersion = os.path.splitext(os.path.basename(notes))[0].split('_')[3]
        # 拼接notes路径
        pathOfNotes = os.path.join(dirOfNotes, notes)
        # 文档对象
        docNotes = Document(pathOfNotes)

        # 获取发布包table的索引号
        # 发布包table的索引号前后并不是一致，但是table的第一行是名固定并且唯一，所以通过第一行确认
        pckTable = ''
        for tab in docNotes.tables:
            firsRow = tab.rows[0]
            if firsRow.cells[0].text == '目录':
                pckTable = tab

        # 某些情况下notes是空的，不做处理
        if pckTable:
            rows = pckTable.rows
            # 若table 行数<1 说明package 列表是写在excel，然后以object形式插入在nots中
            if len(rows) > 1:
                for row in rows:
                    if row.cells[0].text != '目录':
                        info(releaseDate, noteVersion, row.cells[0].text.replace('\n', '').replace('\r',''),row.cells[1].text)
            else:
                # if row is 1, find excel on zip
                tempPath = os.path.join(dirOfNotes, 'temp')
                if not os.path.isdir(tempPath):
                    os.mkdir(tempPath)
                # 临时的zip文件名，将notes复制改名为zip，准备解压提取excel
                zipFilePath = os.path.join(tempPath, os.path.splitext(os.path.basename(notes))[0] + '.zip')
                shutil.copy(pathOfNotes, zipFilePath)

                # 解压zip，提取内部文件列表
                f = zipfile.ZipFile(zipFilePath)

                # 检查notes中是否包含有2个excel，如果有，第二个是package list的文件，否则第一个是
                excelPath = os.path.join(tempPath, 'word/embeddings/Microsoft_Excel_Worksheet1.xlsx')
                for innerFile in f.namelist():
                    #info(innerFile)
                    if 'word/embeddings/Microsoft_Excel_Worksheet2.xlsx' == innerFile:
                        f.extract('word/embeddings/Microsoft_Excel_Worksheet2.xlsx', os.path.join(dirOfNotes, 'temp'))
                        excelPath = os.path.join(tempPath, 'word/embeddings/Microsoft_Excel_Worksheet2.xlsx')
                        break
                # 若worksheet2 文件不存在，那么就默认解压worksheet1
                if not os.path.exists(excelPath):
                    f.extract('word/embeddings/Microsoft_Excel_Worksheet1.xlsx', os.path.join(dirOfNotes, 'temp'))

                # 读取文件内容
                if os.path.exists(excelPath):
                    #info(excelPath)
                    wb = load_workbook(excelPath)
                    ws = wb.active
                    for row in ws.iter_rows():
                        # 当每行数据个数>2, 那么第2，3 列是需要提取的 目录以及文件名
                        if len(row) > 2:
                            info(releaseDate, noteVersion, row[1].value, row[2].value)
                        else:
                            info(releaseDate, noteVersion, row[0].value, row[1].value)

                    wb.close()

                    os.remove(excelPath)

                else:
                    err('Not find excels')

                f.close()

if __name__ == '__main__':
    readNotes()